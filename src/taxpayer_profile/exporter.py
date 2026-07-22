"""Export profiles, abstract trajectories, and update logs to a new workbook."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select

from taxpayer_profile.database import make_engine, make_session_factory
from taxpayer_profile.models import CallerProfile, CallTrajectory, UpdateLog
from taxpayer_profile.profiling import (
    RECEPTION_MODE_CATALOG,
    RECEPTION_MODE_GROUPS,
    classify_reception_mode,
)
from taxpayer_profile.security import PhoneProtector


def _boolean(value: bool | None) -> str:
    if value is None:
        return "无法判断"
    return "是" if value else "否"


def _datetime(value: datetime | None) -> str | None:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else None


def _excel_safe(value: Any) -> Any:
    """Prevent untrusted text from becoming an Excel formula."""

    if isinstance(value, str) and value.lstrip(" \t\r\n").startswith(
        ("=", "+", "-", "@")
    ):
        return "'" + value
    return value


def _safe_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {column: _excel_safe(value) for column, value in row.items()} for row in rows
    ]


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    alternate_fill = PatternFill("solid", fgColor="EAF2F8")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))
    tab_colors = ("1F4E78", "2E75B6", "5B9BD5", "70AD47")
    for sheet_index, worksheet in enumerate(workbook.worksheets):
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_properties.tabColor = tab_colors[sheet_index % len(tab_colors)]
        worksheet.row_dimensions[1].height = 28
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
                if row_index % 2 == 0:
                    cell.fill = alternate_fill
        for column_cells in worksheet.columns:
            letter = column_cells[0].column_letter
            longest = max(
                (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                default=8,
            )
            worksheet.column_dimensions[letter].width = min(max(longest + 2, 12), 52)
    workbook.save(path)


def export_results(
    *, database_path: Path | str, output_path: Path | str, protector: PhoneProtector
) -> Path:
    destination = Path(output_path).expanduser().resolve()
    if destination.exists():
        raise FileExistsError(f"拒绝覆盖已有导出文件：{destination}")

    engine = make_engine(database_path)
    sessions = make_session_factory(engine)
    with sessions() as session:
        profiles = list(
            session.scalars(select(CallerProfile).order_by(CallerProfile.latest_call_time))
        )
        trajectories = list(
            session.scalars(select(CallTrajectory).order_by(CallTrajectory.call_time))
        )
        logs = list(session.scalars(select(UpdateLog).order_by(UpdateLog.started_at)))

    phones = {
        profile.phone_hash: protector.decrypt_phone(profile.phone_encrypted)
        for profile in profiles
    }
    profile_rows: list[dict[str, Any]] = [
        {
            "来电号码": phones[profile.phone_hash],
            "咨询主体": profile.caller_type,
            "细化主体": profile.enterprise_identity,
            "业务熟悉度": profile.proficiency_level,
            "业务熟悉度依据": profile.proficiency_basis,
            "近期情绪状态": profile.emotion_state,
            "情绪状态依据": profile.emotion_basis,
            "首次来电时间": _datetime(profile.first_call_time),
            "最近来电时间": _datetime(profile.latest_call_time),
            "累计来电次数": profile.total_call_count,
            "重复来电次数": profile.repeated_call_count,
            "重复问题次数": profile.repeated_issue_count,
            "未直接解决次数": profile.unresolved_count,
            "工单次数": profile.work_order_count,
            "非正常中断次数": profile.abnormal_end_count,
            "不满次数": profile.dissatisfaction_count,
            "最近核心问题": profile.latest_question,
            "最近专题类别": profile.latest_topic_category,
            "最近需求类别": profile.latest_demand_category,
            "最近登记单位": profile.latest_registration_unit,
            "最近上层问题": profile.latest_father_question,
            "最近是否解决": _boolean(profile.latest_resolved),
            "最近未直接解决原因": profile.latest_unresolved_reason,
            "最近服务效果评估": profile.latest_service_rating,
            "画像摘要": profile.profile_summary,
            "近期问题摘要": profile.recent_questions_summary,
            "未解决问题摘要": profile.unresolved_questions_summary,
            "重复问题摘要": profile.repeated_questions_summary,
            "画像更新时间": _datetime(profile.updated_at),
        }
        for profile in profiles
    ]
    trajectory_rows: list[dict[str, Any]] = [
        {
            "业务编号": item.business_id,
            "转写结果": item.raw_transcript,
            "登记日期": _datetime(item.registration_time),
            "通话开始时间": _datetime(item.raw_call_start_time),
            "通话结束时间": _datetime(item.call_end_time),
            "坐席工号": item.agent_id,
            "坐席姓名": item.agent_name,
            "业务内容": item.business_content,
            "答复内容": item.answer_content,
            "录音路径": item.recording_path,
            "登记单位": item.registration_unit,
            "登记处理方式": item.handling_method,
            "业务类别": item.business_category,
            "来电号码": phones.get(item.phone_hash, ""),
            "满意度": item.satisfaction,
            "呼叫流水号": item.call_serial_number,
            "画像来电时间": _datetime(item.call_time),
            "来电时间来源": item.call_time_source,
            "咨询主体": item.caller_type,
            "细化主体": item.enterprise_identity,
            "核心问题": item.core_question,
            "专题类别": item.topic_category,
            "需求类别": item.demand_category,
            "father_question": item.father_question,
            "father_question_2": item.father_question_2,
            "是否直接解决": _boolean(item.resolved_status),
            "未直接解决原因": item.unresolved_reason,
            "是否工单": _boolean(item.work_order),
            "规则非正常中断": _boolean(item.rule_abnormal_end),
            "语义非正常中断": _boolean(item.model_abnormal_end),
            "是否等待": _boolean(item.waiting_expression),
            "是否潜在推诿": _boolean(item.potential_pushback),
            "是否不满": _boolean(item.taxpayer_dissatisfied),
            "是否联系其他人员或部门": _boolean(item.contacted_other_department),
            "是否主动联系": _boolean(item.active_contacted_other_department),
            "联系对象": item.contact_target,
            "自然问答轮次": item.natural_qa_turns,
            "核心问题轮次": item.core_question_turns,
            "有效问答轮次": item.effective_qa_turns,
            "有效问答内容": item.effective_qa_content,
            "办税熟练程度": item.proficiency_score,
            "熟练程度说明": item.proficiency_summary,
            "业务熟悉度": item.proficiency_level,
            "业务熟悉度依据": item.proficiency_basis,
            "近期情绪状态": item.emotion_state,
            "情绪状态依据": item.emotion_basis,
            "服务效果评估": item.service_rating,
            "服务效果说明": item.service_summary,
            "是否重复来电": _boolean(item.is_repeated_call),
            "上次来电时间": _datetime(item.previous_call_time),
            "距上次来电秒数": item.call_interval,
            "历史累计来电次数": item.historical_call_count,
            "是否重复问题": _boolean(item.is_repeated_issue),
            "匹配历史业务编号": item.matched_previous_business_id,
            "匹配历史核心问题": item.matched_previous_question,
            "匹配历史咨询时间": _datetime(item.matched_previous_call_time),
            "前次是否直接解决": _boolean(item.previous_issue_resolved),
            "重复咨询原因": item.repeat_reason,
            "重复问题说明": item.repeat_summary,
            "本地候选相似度": item.repeat_candidate_score,
            "重复问题置信度": item.repeat_confidence,
            "重复问题判断状态": item.repeat_review_status,
            "分析状态": item.analysis_status,
            "分析版本": item.analysis_version,
            "输入模式": item.input_mode,
            "分析来源": item.analysis_source,
            "模型名称": item.model_name,
            "提示词版本": item.prompt_version,
            "抽取版本": item.extraction_version,
            "身份来源": item.enterprise_identity_source,
            "身份冲突": _boolean(item.enterprise_identity_conflict),
            "来源文件": item.source_filename,
        }
        for item in trajectories
    ]
    log_rows: list[dict[str, Any]] = [
        {
            "批次编号": item.batch_id,
            "数据日期": item.data_date,
            "输入文件名": item.input_filename,
            "输入文件指纹": item.input_fingerprint,
            "输入模式": item.input_mode,
            "分析版本": item.analysis_version,
            "源数据行数": item.source_row_count,
            "开始时间": _datetime(item.started_at),
            "结束时间": _datetime(item.finished_at),
            "新增来电": item.new_call_count,
            "新增号码": item.new_phone_count,
            "更新画像": item.updated_profile_count,
            "重复来电": item.repeated_call_count,
            "重复问题": item.repeated_issue_count,
            "未直接解决": item.unresolved_count,
            "失败记录": item.failed_count,
            "跳过记录": item.skipped_count,
            "冲突记录": item.conflict_count,
            "状态": item.status,
            "摘要": item.summary,
        }
        for item in logs
    ]

    destination.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(destination, engine="openpyxl") as writer:
        pd.DataFrame(_safe_rows(profile_rows)).to_excel(
            writer, sheet_name="号码画像", index=False
        )
        pd.DataFrame(_safe_rows(trajectory_rows)).to_excel(
            writer, sheet_name="来电轨迹", index=False
        )
        pd.DataFrame(_safe_rows(log_rows)).to_excel(
            writer, sheet_name="更新摘要", index=False
        )
    _format_workbook(destination)
    return destination


def export_profile_rule_workbooks(output_directory: Path | str) -> tuple[Path, Path]:
    """Write external-facing methodology and service-guidance workbooks."""

    directory = Path(output_directory).expanduser().resolve()
    directory.mkdir(parents=True, exist_ok=True)
    mapping_path = directory / "纳税人特征与接待模式映射表.xlsx"
    guidance_path = directory / "接待模式服务建议表.xlsx"

    mapping_notes = [
        {
            "说明项": "文档用途",
            "内容": "说明纳税人特征如何分别推导情绪响应、事项承接和表达方式，并组合形成完整接待策略；供方案评审、演示和规则核验使用。",
        },
        {
            "说明项": "输入结构",
            "内容": "输入由业务熟悉度、近期情绪状态和五项历史服务事实组成，三类信息共同描述纳税人本身情况及近期服务经历。",
        },
        {
            "说明项": "统计窗口",
            "内容": "以最近一次来电日期为锚点，取最近五个工作日。业务熟悉度和情绪采用窗口内最近一次有效判断；历史服务事实统计窗口内发生次数。",
        },
        {
            "说明项": "推导方式",
            "内容": "三个类别分别判断、同时生效：情绪响应三选一、事项承接二选一、表达方式三选一；不再用单一全局优先级覆盖其他维度。",
        },
        {
            "说明项": "识别与映射",
            "内容": "特征可复用既有分析字段，也可由新增来电的大模型分析结果提供；从特征到三个分项模式均采用确定性规则映射。",
        },
        {
            "说明项": "重要边界",
            "内容": "类别内结果唯一：情绪响应按安抚修复、稳定预期、平稳接待依次判断；类别之间无覆盖关系，例如专业且不满可同时得到安抚修复和结论直述。",
        },
        {
            "说明项": "使用边界",
            "内容": "组合策略提供宏观沟通方向。坐席仍需结合本次诉求核验业务事实、适用政策和具体办理口径。",
        },
    ]

    field_rows = [
        {
            "维度": "业务熟悉度",
            "取值或事实": "专业",
            "判定口径": "能准确描述业务场景、条件或办理节点，理解相关术语，追问集中在边界或结果",
            "聚合方式": "最近五个工作日内最近一次有效判断",
            "映射作用": "表达方式选择结论直述",
        },
        {
            "维度": "业务熟悉度",
            "取值或事实": "了解",
            "判定口径": "知道事项和基本办理方向，能够说明主要问题，但仍需确认部分规则、材料或节点",
            "聚合方式": "最近五个工作日内最近一次有效判断",
            "映射作用": "表达方式选择重点解释",
        },
        {
            "维度": "业务熟悉度",
            "取值或事实": "小白",
            "判定口径": "对基础概念、办理入口或操作路径明显不熟悉，需要通俗解释和分段引导",
            "聚合方式": "最近五个工作日内最近一次有效判断",
            "映射作用": "表达方式选择通俗引导",
        },
        {
            "维度": "近期情绪状态",
            "取值或事实": "平稳",
            "判定口径": "表达基本有序，没有明显担忧、责备或负面评价",
            "聚合方式": "最近五个工作日内最近一次有效判断",
            "映射作用": "未出现修复信号时，情绪响应选择平稳接待",
        },
        {
            "维度": "近期情绪状态",
            "取值或事实": "焦虑",
            "判定口径": "担忧时限、结果、处罚或损失，反复追问进度，但没有明确负面评价",
            "聚合方式": "最近五个工作日内最近一次有效判断",
            "映射作用": "未出现修复信号时，情绪响应选择稳定预期",
        },
        {
            "维度": "近期情绪状态",
            "取值或事实": "不满",
            "判定口径": "出现明确负面评价、责备、拒绝或投诉倾向；不限定不满对象",
            "聚合方式": "最近五个工作日内最近一次有效判断",
            "映射作用": "情绪响应选择安抚修复",
        },
        {
            "维度": "历史服务事实",
            "取值或事实": "等待推诿",
            "判定口径": "存在让纳税人等待表述=是，且坐席存在潜在推诿行为=是",
            "聚合方式": "统计最近五个工作日内命中次数",
            "映射作用": "发生至少一次，情绪响应选择安抚修复",
        },
        {
            "维度": "历史服务事实",
            "取值或事实": "历史工单",
            "判定口径": "是否工单=是",
            "聚合方式": "统计最近五个工作日内命中次数",
            "映射作用": "发生至少一次，事项承接选择历史跟进",
        },
        {
            "维度": "历史服务事实",
            "取值或事实": "异常中断",
            "判定口径": "既有分析采用已确认字段；新增来电仅在明确中断、挂断或语义判断命中时计入，缺少结束语不单独认定",
            "聚合方式": "规则判断或语义判断任一为是即计入；统计最近五个工作日内命中次数",
            "映射作用": "发生至少一次，事项承接选择历史跟进",
        },
        {
            "维度": "历史服务事实",
            "取值或事实": "联系后未解决",
            "判定口径": "对话中存在联系相关人员或部门=是，且坐席是否解决纳税人问题=否",
            "聚合方式": "统计最近五个工作日内同时满足条件的次数",
            "映射作用": "发生至少一次，事项承接选择历史跟进",
        },
        {
            "维度": "历史服务事实",
            "取值或事实": "对坐席不满",
            "判定口径": "纳税人是否对当前坐席或本通热线存在不满=是",
            "聚合方式": "统计最近五个工作日内命中次数",
            "映射作用": "发生至少一次，情绪响应选择安抚修复",
        },
    ]
    mapping_rows = [
        {
            "模式类别": group["label"],
            "类别说明": group["description"],
            "类别内判断顺序": index,
            "分项模式": mode["label"],
            "触发条件": mode["rule"],
            "接待重点": mode["focus"],
            "组合关系": "从本类别选择一项；与另外两个类别的结果同时生效",
        }
        for group in RECEPTION_MODE_GROUPS
        for index, mode in enumerate(group["modes"], 1)
    ]

    example_specs = [
        ("E01", "专业", "平稳", {}, "平稳→平稳接待；无待衔接事实→诉求确认；专业→结论直述"),
        ("E02", "了解", "焦虑", {}, "焦虑→稳定预期；无待衔接事实→诉求确认；了解→重点解释"),
        ("E03", "小白", "平稳", {}, "平稳→平稳接待；无待衔接事实→诉求确认；小白→通俗引导"),
        ("E04", "暂无法判断", "焦虑", {}, "焦虑→稳定预期；无待衔接事实→诉求确认；熟悉度证据不足→通俗引导"),
        ("E05", "专业", "不满", {}, "不满→安抚修复；无待衔接事实→诉求确认；专业→结论直述"),
        (
            "E06",
            "专业",
            "平稳",
            {"wait_pushback_count": 1},
            "等待推诿→安抚修复；无待衔接事实→诉求确认；专业→结论直述",
        ),
        (
            "E07",
            "了解",
            "平稳",
            {"dissatisfaction_count": 1},
            "对坐席不满→安抚修复；无待衔接事实→诉求确认；了解→重点解释",
        ),
        (
            "E08",
            "专业",
            "平稳",
            {"work_order_count": 1},
            "平稳→平稳接待；历史工单→历史跟进；专业→结论直述",
        ),
        (
            "E09",
            "小白",
            "平稳",
            {"abnormal_end_count": 1},
            "平稳→平稳接待；异常中断→历史跟进；小白→通俗引导",
        ),
        (
            "E10",
            "了解",
            "焦虑",
            {"contact_unresolved_count": 1},
            "焦虑→稳定预期；联系后未解决→历史跟进；了解→重点解释",
        ),
        (
            "E11",
            "专业",
            "平稳",
            {"wait_pushback_count": 1, "work_order_count": 1},
            "等待推诿→安抚修复；历史工单→历史跟进；专业→结论直述，三个结果并行保留",
        ),
        (
            "E12",
            "了解",
            "不满",
            {"work_order_count": 1, "abnormal_end_count": 1},
            "不满→安抚修复；工单或异常中断→历史跟进；了解→重点解释，三个结果并行保留",
        ),
    ]

    fact_labels = {
        "wait_pushback_count": "等待推诿",
        "work_order_count": "历史工单",
        "abnormal_end_count": "异常中断",
        "contact_unresolved_count": "联系后未解决",
        "dissatisfaction_count": "对坐席不满",
    }
    example_rows: list[dict[str, Any]] = []
    for example_id, proficiency, emotion, facts, derivation in example_specs:
        result = classify_reception_mode(
            proficiency_level=proficiency,
            emotion_state=emotion,
            **facts,
        )
        rendered_facts = "、".join(
            f"{fact_labels[key]}{count}次" for key, count in facts.items() if count
        )
        components = {component.category_id: component for component in result.components}
        example_rows.append(
            {
                "示例编号": example_id,
                "业务熟悉度": proficiency,
                "近期情绪": emotion,
                "历史事实（最近五个工作日）": rendered_facts or "无",
                "规则推导": derivation,
                "情绪响应": components["emotion_response"].mode,
                "事项承接": components["matter_continuity"].mode,
                "表达方式": components["information_delivery"].mode,
                "组合接待策略": result.mode,
                "接待重点": result.focus,
                "需要避免": result.avoid,
            }
        )

    combination_rows = []
    combination_index = 0
    for emotion_mode in RECEPTION_MODE_GROUPS[0]["modes"]:
        for continuity_mode in RECEPTION_MODE_GROUPS[1]["modes"]:
            for expression_mode in RECEPTION_MODE_GROUPS[2]["modes"]:
                combination_index += 1
                selected = (emotion_mode, continuity_mode, expression_mode)
                combination_rows.append(
                    {
                        "组合编号": f"C{combination_index:02d}",
                        "情绪响应": emotion_mode["label"],
                        "事项承接": continuity_mode["label"],
                        "表达方式": expression_mode["label"],
                        "组合接待策略": " · ".join(
                            str(item["label"]) for item in selected
                        ),
                        "综合服务方向": "；".join(
                            str(item["focus"]) for item in selected
                        ),
                    }
                )

    with pd.ExcelWriter(mapping_path, engine="openpyxl") as writer:
        pd.DataFrame(_safe_rows(mapping_notes)).to_excel(
            writer, sheet_name="使用说明", index=False
        )
        pd.DataFrame(_safe_rows(mapping_rows)).to_excel(
            writer, sheet_name="模式映射规则", index=False
        )
        pd.DataFrame(_safe_rows(field_rows)).to_excel(
            writer, sheet_name="字段判定口径", index=False
        )
        pd.DataFrame(_safe_rows(example_rows)).to_excel(
            writer, sheet_name="推导示例", index=False
        )
        pd.DataFrame(_safe_rows(combination_rows)).to_excel(
            writer, sheet_name="18种组合总览", index=False
        )
    _format_workbook(mapping_path)

    mode_details = {
        "安抚修复": {
            "建议开场方向": "先确认已了解其关注点和既往体验，再说明本通将优先核实的事项。",
            "结束前确认": "确认已说明当前结论、后续责任节点和可追踪方式。",
        },
        "稳定预期": {
            "建议开场方向": "先确认其最关注的时限、结果或影响，再说明当前可确认的范围。",
            "结束前确认": "确认其已理解当前结论、时间预期和仍需等待的节点。",
        },
        "平稳接待": {
            "建议开场方向": "自然确认本次诉求和当前卡点，按正常节奏进入事项处理。",
            "结束前确认": "确认本次核心问题和后续安排已说明清楚。",
        },
        "历史跟进": {
            "建议开场方向": "先确认本次是否延续历史事项，并核对业务编号、工单或最近处理节点。",
            "结束前确认": "确认当前进展、下一责任节点、所需补充信息及合理反馈预期。",
        },
        "诉求确认": {
            "建议开场方向": "先确认本次来电主体、事项和当前卡点，不预设其延续历史问题。",
            "结束前确认": "确认本次实际诉求已经得到对应说明。",
        },
        "结论直述": {
            "建议开场方向": "先回应核心判断，再说明适用条件和必要办理节点。",
            "结束前确认": "确认来电人已掌握结论、边界条件及仍需自行核验的事项。",
        },
        "重点解释": {
            "建议开场方向": "先说明关键判断，再解释必要条件和容易混淆的节点。",
            "结束前确认": "确认关键条件和主要差异已被理解。",
        },
        "通俗引导": {
            "建议开场方向": "先用简明语言说明事项目标，再从当前最关键的一个节点开始解释。",
            "结束前确认": "确认来电人理解当前节点，并知道下一步应准备或操作的内容。",
        },
    }
    guidance_rows = [
        {
            "模式类别": mode["category"],
            "分项模式": mode["label"],
            "适用情形": mode["rule"],
            "接待重点": mode["focus"],
            "沟通方式": mode["communication"],
            "建议开场方向": mode_details[str(mode["label"])]["建议开场方向"],
            "结束前确认": mode_details[str(mode["label"])]["结束前确认"],
            "需要避免": mode["avoid"],
        }
        for mode in RECEPTION_MODE_CATALOG
    ]

    guidance_notes = [
        {
            "说明项": "文档用途",
            "内容": "说明三类、八项分项模式的服务目标和沟通方向，以及18种组合策略的应用方式；供坐席参考、方案展示和培训说明。",
        },
        {
            "说明项": "使用顺序",
            "内容": "先根据《纳税人特征与接待模式映射表》从三个类别中各选择一项，再同时落实三个分项模式的服务重点。",
        },
        {
            "说明项": "动态调整",
            "内容": "组合策略来自历史画像。接听过程中如出现新的情绪、事项或理解程度证据，坐席应以本次实际情况动态调整相应分项方式。",
        },
        {
            "说明项": "建议边界",
            "内容": "表内内容为宏观接待建议，不生成具体业务结论，不替代政策核验、身份核验及规范服务要求。",
        },
    ]

    scenario_indexes = {"E02", "E03", "E05", "E06", "E07", "E08", "E09", "E10", "E11"}
    scenarios: list[dict[str, Any]] = []
    for row in example_rows:
        if row["示例编号"] not in scenario_indexes:
            continue
        scenarios.append(
            {
                "场景编号": row["示例编号"],
                "特征组合": (
                    f"{row['业务熟悉度']}；{row['近期情绪']}；"
                    f"{row['历史事实（最近五个工作日）']}"
                ),
                "情绪响应": row["情绪响应"],
                "事项承接": row["事项承接"],
                "表达方式": row["表达方式"],
                "组合接待策略": row["组合接待策略"],
                "推导依据": row["规则推导"],
                "现场侧重点": row["接待重点"],
                "需要避免": row["需要避免"],
            }
        )

    with pd.ExcelWriter(guidance_path, engine="openpyxl") as writer:
        pd.DataFrame(_safe_rows(guidance_notes)).to_excel(
            writer, sheet_name="使用说明", index=False
        )
        pd.DataFrame(_safe_rows(guidance_rows)).to_excel(
            writer, sheet_name="接待模式服务建议", index=False
        )
        pd.DataFrame(_safe_rows(combination_rows)).to_excel(
            writer, sheet_name="18种组合总览", index=False
        )
        pd.DataFrame(_safe_rows(scenarios)).to_excel(
            writer, sheet_name="典型服务场景", index=False
        )
    _format_workbook(guidance_path)
    return mapping_path, guidance_path
