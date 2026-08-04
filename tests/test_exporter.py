from collections import Counter
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

from cryptography.fernet import Fernet
from openpyxl import load_workbook

from taxpayer_profile.application.dashboard_service import DashboardService
from taxpayer_profile.application.web_dto import frequent_then_unresolved_rate_rows
from taxpayer_profile.application.history_service import HistoryService
from taxpayer_profile.application.profile_showcase_service import ProfileShowcaseService
from taxpayer_profile.application.web_dto import unresolved_rate_rows
from taxpayer_profile.config import Settings
from taxpayer_profile.database import (
    create_schema,
    make_engine,
    make_session_factory,
    transactional_session,
)
from taxpayer_profile.exporter import export_profile_rule_workbooks, export_results
from taxpayer_profile.models import CallerProfile, CallTrajectory, UpdateLog
from taxpayer_profile.query import (
    _agent_answer_summary,
    _five_workdays,
    _recent_workday_statistics,
    query_profile,
)
from taxpayer_profile.security import PhoneProtector
from taxpayer_profile.web_app import DemoService, _district_unit_label, _segmented_rows


def _seed_database(path: Path, protector: PhoneProtector) -> None:
    engine = make_engine(path)
    create_schema(engine)
    sessions = make_session_factory(engine)
    phone_hash = protector.hash_phone("13800000001")
    call_time = datetime(2026, 6, 1, 9, 0)
    with transactional_session(sessions) as session:
        session.add(
            CallerProfile(
                phone_hash=phone_hash,
                phone_encrypted=protector.encrypt_phone("13800000001"),
                caller_type="个人",
                enterprise_identity="不适用",
                first_call_time=call_time,
                latest_call_time=call_time,
                total_call_count=1,
                latest_business_id="BIZ-1",
                latest_question='=HYPERLINK("https://example.invalid","申报问题")',
                latest_resolved=True,
                latest_service_rating="良好",
            )
        )
        session.add(
            CallTrajectory(
                business_id="BIZ-1",
                phone_hash=phone_hash,
                call_time=call_time,
                registration_time=call_time,
                raw_call_start_time=call_time,
                raw_phone_encrypted=protector.encrypt_phone("13800000001"),
                raw_transcript="纳税人咨询申报问题，坐席给出办理路径。",
                business_content="咨询申报问题",
                answer_content="已告知办理路径",
                agent_answer_summary="坐席说明了申报办理路径。",
                registration_unit="第一税务所",
                caller_type="个人",
                enterprise_identity="不适用",
                core_question='=HYPERLINK("https://example.invalid","申报问题")',
                resolved_status=True,
                work_order=False,
                rule_abnormal_end=False,
                is_repeated_call=False,
                historical_call_count=1,
                is_repeated_issue=False,
                service_rating="良好",
                analysis_status="completed",
                analysis_version="test-v1",
            )
        )
        session.add(
            UpdateLog(
                batch_id="batch-1",
                data_date="2026-06-01..2026-06-09",
                input_filename="synthetic.xlsx",
                started_at=call_time,
                finished_at=call_time,
                new_call_count=1,
                new_phone_count=1,
                updated_profile_count=1,
                repeated_call_count=0,
                repeated_issue_count=0,
                unresolved_count=0,
                failed_count=0,
                status="completed",
            )
        )


def test_export_has_three_formatted_worksheets_and_query_interface(
    tmp_path: Path,
) -> None:
    database = tmp_path / "profiles.sqlite3"
    output = tmp_path / "results.xlsx"
    protector = PhoneProtector("test-hash-key", Fernet.generate_key().decode())
    _seed_database(database, protector)

    export_results(database_path=database, output_path=output, protector=protector)
    workbook = load_workbook(output)

    assert workbook.sheetnames == ["号码画像", "来电轨迹", "更新摘要"]
    for sheet in workbook.worksheets:
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref is not None
    assert workbook["号码画像"]["A2"].value == "13800000001"
    trajectory_sheet = workbook["来电轨迹"]
    trajectory_headers = [cell.value for cell in trajectory_sheet[1]]
    assert trajectory_headers[:16] == [
        "业务编号",
        "转写结果",
        "登记日期",
        "通话开始时间",
        "通话结束时间",
        "坐席工号",
        "坐席姓名",
        "业务内容",
        "答复内容",
        "录音路径",
        "登记单位",
        "登记处理方式",
        "业务类别",
        "来电号码",
        "满意度",
        "呼叫流水号",
    ]
    resolved_column = trajectory_headers.index("是否直接解决") + 1
    assert trajectory_sheet.cell(row=2, column=resolved_column).value == "是"
    answer_summary_column = trajectory_headers.index("坐席答复提炼") + 1
    assert trajectory_sheet.cell(row=2, column=answer_summary_column).value == (
        "坐席说明了申报办理路径。"
    )
    profile_sheet = workbook["号码画像"]
    profile_headers = [cell.value for cell in profile_sheet[1]]
    latest_question_column = profile_headers.index("最近核心问题") + 1
    latest_question_cell = profile_sheet.cell(row=2, column=latest_question_column)
    assert latest_question_cell.data_type == "s"
    assert latest_question_cell.value.startswith("'=")
    question_column = trajectory_headers.index("核心问题") + 1
    question_cell = trajectory_sheet.cell(row=2, column=question_column)
    assert question_cell.data_type == "s"
    assert question_cell.value.startswith("'=")

    result = query_profile(
        phone="13800000001", database_path=database, protector=protector
    )
    assert result is not None
    assert "phone" not in result
    assert result["total_call_count"] == 1
    assert result["agent_context"]["statistics"]["total_calls"] == 1
    assert result["latest_agent_answer"] == "坐席说明了申报办理路径。"
    assert "知识库暂未接入" in result["standard_answer"]
    assert result["agent_context"]["recommended_mode"] == (
        "通俗引导 · 平稳接待 · 当前诉求确认"
    )
    assert len(result["agent_context"]["recommended_modes"]) == 3
    assert "service_suggestion" not in result
    assert result["trajectories"][0]["business_id"] == "BIZ-1"
    assert result["trajectories"][0]["work_order"] is False
    assert "matched_previous_question" in result["trajectories"][0]


def test_profile_rule_workbooks_are_self_contained_and_cover_derivations(
    tmp_path: Path,
) -> None:
    mapping_path, guidance_path = export_profile_rule_workbooks(tmp_path)

    mapping = load_workbook(mapping_path)
    assert mapping.sheetnames == [
        "使用说明",
        "模式映射规则",
        "字段判定口径",
        "推导示例",
        "18种组合总览",
    ]
    assert mapping["模式映射规则"].max_row == 9
    assert mapping["字段判定口径"].max_row == 12
    assert mapping["推导示例"].max_row == 13
    assert mapping["18种组合总览"].max_row == 19
    assert {cell.value for cell in mapping["推导示例"]["G"][1:]} == {
        "安抚修复",
        "稳定预期",
        "平稳接待",
    }
    assert {cell.value for cell in mapping["推导示例"]["H"][1:]} == {
        "历史诉求跟进",
        "当前诉求确认",
    }
    assert {cell.value for cell in mapping["推导示例"]["F"][1:]} == {
        "结论直述",
        "重点解释",
        "通俗引导",
    }
    mapping_text = " ".join(
        str(cell.value or "")
        for sheet in mapping.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    assert "缺少结束语不单独认定" in mapping_text
    assert "号码级" not in mapping_text

    guidance = load_workbook(guidance_path)
    assert guidance.sheetnames == [
        "使用说明",
        "接待模式服务建议",
        "18种组合总览",
        "典型服务场景",
    ]
    assert guidance["接待模式服务建议"].max_row == 9
    assert guidance["18种组合总览"].max_row == 19
    assert guidance["典型服务场景"].max_row == 10
    guidance_text = " ".join(
        str(cell.value or "")
        for sheet in guidance.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    assert "宏观接待建议" in guidance_text
    assert "号码级" not in guidance_text


def test_topic_distribution_selects_unresolved_top_five_from_ten_most_frequent() -> None:
    counts = Counter({
        "专题01": 30,
        "专题02": 29,
        "专题03": 28,
        "专题04": 27,
        "专题05": 26,
        "专题06": 25,
        "专题07": 24,
        "专题08": 23,
        "专题09": 22,
        "专题10": 21,
        "低频高未解决": 2,
    })
    resolution = {
        label: Counter({"resolved": value - index, "unresolved": index})
        for index, (label, value) in enumerate(counts.items(), start=1)
    }
    resolution["低频高未解决"] = Counter({"unresolved": 2})

    rows = frequent_then_unresolved_rate_rows(counts, resolution)

    assert [row["label"] for row in rows] == [
        "专题10", "专题09", "专题08", "专题07", "专题06"
    ]
    assert "低频高未解决" not in {row["label"] for row in rows}


def test_web_dashboard_and_history_are_read_only_and_mask_phone(
    tmp_path: Path,
) -> None:
    database = tmp_path / "profiles.sqlite3"
    protector = PhoneProtector("test-hash-key", Fernet.generate_key().decode())
    _seed_database(database, protector)
    settings = Settings(
        database_path=database,
        llm_base_url=None,
        llm_api_key=None,
        llm_model=None,
        phone_hash_key="test-hash-key",
        phone_encryption_key=None,
    )
    service = DemoService(database, protector, settings)

    dashboard = service.dashboard_summary()
    assert dashboard["overview"]["total_profiles"] == 1
    assert dashboard["overview"]["total_calls"] == 1
    assert dashboard["overview"]["resolved_rate"] == 100.0
    assert dashboard["caller_types"] == [{"label": "个人", "value": 1}]
    assert len(dashboard["daily_calls"]) == 14
    assert dashboard["daily_calls"][-1]["value"] == 1
    assert dashboard["resolution_status"] == [
        {"label": "已直接解决", "value": 1}
    ]
    assert dashboard["question_categories"] == []
    assert dashboard["personal_resolution"] == [
        {"label": "已直接解决", "value": 1}
    ]
    assert dashboard["enterprise_resolution"] == []
    assert dashboard["enterprise_identity_resolution_rates"] == []
    assert dashboard["registration_unit_resolution"][0]["label"] == "第一税务所"
    assert dashboard["registration_unit_resolution"][0]["resolved_rate"] == 100.0
    assert dashboard["caller_resolution_rates"] == [
        {
            "label": "个人",
            "resolved": 1,
            "unresolved": 0,
            "unknown": 0,
            "eligible_total": 1,
            "resolved_rate": 100.0,
        },
        {
            "label": "企业",
            "resolved": 0,
            "unresolved": 0,
            "unknown": 0,
            "eligible_total": 0,
            "resolved_rate": None,
        },
    ]
    assert dashboard["historical_facts"][1]["label"] == "存在联系相关部门或人员且未解决"
    assert [item["label"] for item in dashboard["historical_facts"]] == [
        "历史工单",
        "存在联系相关部门或人员且未解决",
        "异常中断",
        "等待推诿",
        "对坐席不满",
    ]
    assert dashboard["unresolved_question_hotspots"] == {
        "all": [], "personal": [], "enterprise": []
    }
    assert dashboard["unresolved_distributions"]["topics"] == []
    assert "data_quality" not in dashboard
    assert dashboard["service_signals"][0]["value"] == 0
    assert dashboard["latest_update"]["input_filename"] == "synthetic.xlsx"

    history = service.history_page(page=1, page_size=10, phone="13800000001")
    assert history["filtered"] is True
    assert history["total"] == 1
    assert history["total_pages"] == 1
    assert history["items"][0]["masked_phone"] == "138****0001"
    assert "13800000001" not in str(history)

    detail = service.history_detail("BIZ-1")
    assert detail is not None
    assert detail["original"]["masked_phone"] == "138****0001"
    assert detail["original"]["transcript"] == "纳税人咨询申报问题，坐席给出办理路径。"
    assert detail["original"]["registration_unit"] == "第一税务所"
    assert detail["extracted"]["resolved"] is True
    assert detail["extracted"]["repeat_review_status"] == "not_required"
    assert "13800000001" not in str(detail)

    catalog = service.profile_showcase_catalog()
    assert len(catalog["items"]) == 1
    assert catalog["items"][0]["masked_phone"] == "138****0001"
    assert catalog["summary"]["dimension_count"] == 3
    assert catalog["summary"]["fact_count"] == 5
    assert catalog["summary"]["mode_group_count"] == 3
    assert catalog["summary"]["mode_count"] == 8
    assert "13800000001" not in str(catalog)
    showcase = service.profile_showcase(profile_key=catalog["items"][0]["profile_key"])
    assert len(showcase["before"]["result"]["mode_components"]) == 3
    assert len(showcase["before"]["profile_model"]["items"]) == 3
    assert showcase["before"]["profile_model"]["active_category_count"] >= 3
    assert showcase["derivation_evidence"]["caller"]["masked_phone"] == "138****0001"
    assert "basis" in showcase["derivation_evidence"]["proficiency"]
    assert "basis" in showcase["derivation_evidence"]["emotion"]
    assert "events" in showcase["derivation_evidence"]["facts"]
    assert showcase["derivation_evidence"]["proficiency"]["source"]["business_id"] == "BIZ-1"
    assert "after" not in showcase
    assert "scenario" not in showcase
    assert "13800000001" not in str(showcase)


def test_dashboard_hotspot_replaces_extraction_missing_value_boilerplate() -> None:
    assert DashboardService._hotspot_question_label("nan") is None
    assert DashboardService._hotspot_question_label("['问题待归类']") is None
    assert DashboardService._hotspot_question_label(
        "提供的问题组中均为“nan”，无法提炼出具体的核心问题。"
    ) is None
    assert DashboardService._hotspot_question_label(
        "这些内容都是“我无法给到相关内容”的重复，无法提炼出核心问题。"
    ) is None


def test_topic_rate_rows_exclude_placeholder_categories() -> None:
    rows = unresolved_rate_rows(
        Counter({"暂未分类": 3, "其他": 2, "['其他']": 2, "未提取出标签": 2, "申报缴税": 1}),
        {
            "暂未分类": Counter({"unresolved": 3}),
            "其他": Counter({"unresolved": 2}),
            "['其他']": Counter({"unresolved": 2}),
            "未提取出标签": Counter({"unresolved": 2}),
            "申报缴税": Counter({"unresolved": 1}),
        },
        exclude_other=True,
        exclude_unclassified=True,
    )
    assert [row["label"] for row in rows] == ["申报缴税"]


def test_web_facade_composes_independent_read_use_cases(tmp_path: Path) -> None:
    """The HTTP-facing façade must stay a delegator as read use cases evolve."""

    database = tmp_path / "profiles.sqlite3"
    protector = PhoneProtector("test-hash-key", Fernet.generate_key().decode())
    _seed_database(database, protector)
    service = DemoService(
        database,
        protector,
        Settings(
            database_path=database,
            llm_base_url=None,
            llm_api_key=None,
            llm_model=None,
            phone_hash_key="test-hash-key",
            phone_encryption_key=None,
        ),
    )

    assert isinstance(service.dashboard, DashboardService)
    assert isinstance(service.history, HistoryService)
    assert isinstance(service.showcase, ProfileShowcaseService)
    assert service.profile_advice.sessions is service._sessions
    assert service.dashboard_summary() == service.dashboard.summary()
    assert service.history_page(phone="13800000001") == service.history.page(
        phone="13800000001"
    )


def test_agent_answer_summary_never_falls_back_to_local_text_rules() -> None:
    pending = SimpleNamespace(
        agent_answer_summary=None,
        model_name=None,
        answer_content="本地原始答复内容",
        raw_transcript="坐席：本地转写内容",
    )

    assert _agent_answer_summary(pending) == "该通记录尚未完成坐席答复的模型提炼。"


def test_dashboard_top_five_skips_other_and_backfills_next_category() -> None:
    counts = Counter(
        {"其他": 20, "类别一": 9, "类别二": 8, "类别三": 7, "类别四": 6, "类别五": 5, "类别六": 4}
    )
    resolution = {
        label: Counter({"resolved": value}) for label, value in counts.items()
    }

    rows = _segmented_rows(counts, resolution, limit=5, exclude_other=True)

    assert [row["label"] for row in rows] == [
        "类别一",
        "类别二",
        "类别三",
        "类别四",
        "类别五",
    ]
    assert rows[0]["share"] == round(9 * 100 / sum(counts.values()), 1)


def test_registration_unit_visual_labels_use_required_short_names() -> None:
    expected = {
        "上海市税务局": "中心",
        "第三税务分局": "三分局",
        "自贸区分局": "自贸区",
        "浦东新区税务局": "浦东",
        "奉贤区税务局": "奉贤",
        "闵行区税务局": "闵行",
        "宝山区税务局": "宝山",
        "金山区税务局": "金山",
        "长宁区税务局": "长宁",
        "崇明区税务局": "崇明",
        "普陀区税务局": "普陀",
        "杨浦区税务局": "杨浦",
        "静安区税务局": "静安",
        "松江区税务局": "松江",
        "嘉定区税务局": "嘉定",
        "青浦区税务局": "青浦",
        "徐汇区税务局": "徐汇",
        "虹口区税务局": "虹口",
        "黄浦区税务局": "黄浦",
    }

    assert {
        source: _district_unit_label(source) for source in expected
    } == expected
    assert _district_unit_label("第一税务所") == "第一税务所"


def test_showcase_catalog_returns_five_defaults_and_searches_all_profiles(
    tmp_path: Path,
) -> None:
    database = tmp_path / "profiles.sqlite3"
    protector = PhoneProtector("test-hash-key", Fernet.generate_key().decode())
    engine = make_engine(database)
    create_schema(engine)
    sessions = make_session_factory(engine)
    with transactional_session(sessions) as session:
        for index in range(7):
            phone = f"138000000{index + 1:02d}"
            session.add(
                CallerProfile(
                    phone_hash=protector.hash_phone(phone),
                    phone_encrypted=protector.encrypt_phone(phone),
                    first_call_time=datetime(2026, 6, index + 1, 9),
                    latest_call_time=datetime(2026, 6, index + 1, 9),
                    total_call_count=1,
                    proficiency_level="了解",
                    emotion_state="平稳",
                )
            )
    settings = Settings(
        database_path=database,
        llm_base_url=None,
        llm_api_key=None,
        llm_model=None,
        phone_hash_key="test-hash-key",
        phone_encryption_key=None,
    )
    service = DemoService(database, protector, settings)

    default_catalog = service.profile_showcase_catalog()
    searched_catalog = service.profile_showcase_catalog(query="13800000001")

    assert len(default_catalog["items"]) == 5
    assert default_catalog["summary"]["profile_count"] == 7
    assert len(searched_catalog["items"]) == 1
    assert searched_catalog["items"][0]["masked_phone"] == "138****0001"
    assert "13800000001" not in str(searched_catalog)


def test_showcase_catalog_reflects_profiles_added_after_first_read(
    tmp_path: Path,
) -> None:
    database = tmp_path / "profiles.sqlite3"
    protector = PhoneProtector("test-hash-key", Fernet.generate_key().decode())
    engine = make_engine(database)
    create_schema(engine)
    sessions = make_session_factory(engine)

    def add_profile(phone: str, when: datetime) -> None:
        with transactional_session(sessions) as session:
            session.add(
                CallerProfile(
                    phone_hash=protector.hash_phone(phone),
                    phone_encrypted=protector.encrypt_phone(phone),
                    first_call_time=when,
                    latest_call_time=when,
                    total_call_count=1,
                    proficiency_level="了解",
                    emotion_state="平稳",
                )
            )

    add_profile("13800000001", datetime(2026, 6, 1, 9))
    service = ProfileShowcaseService(sessions, protector)
    assert service.catalog()["summary"]["profile_count"] == 1

    add_profile("13800000002", datetime(2026, 6, 2, 9))
    refreshed = service.catalog()

    assert refreshed["summary"]["profile_count"] == 2
    assert [item["masked_phone"] for item in refreshed["items"]] == [
        "138****0002",
        "138****0001",
    ]


def test_information_overview_uses_five_statutory_service_days() -> None:
    assert [item.isoformat() for item in _five_workdays(datetime(2026, 6, 22).date())] == [
        "2026-06-15",
        "2026-06-16",
        "2026-06-17",
        "2026-06-18",
        "2026-06-22",
    ]
    trajectories = [
        SimpleNamespace(
            call_time=datetime(2026, 6, 22, 9),
            demand_category="政策咨询类, 操作辅导类",
            topic_category="个税汇算",
            work_order=False,
            resolved_status=True,
            taxpayer_dissatisfied=False,
            is_repeated_issue=False,
            repeat_label_expires_at=None,
        ),
        SimpleNamespace(
            call_time=datetime(2026, 6, 21, 9),
            demand_category="政策咨询类",
            topic_category="个税汇算",
            work_order=True,
            resolved_status=False,
            taxpayer_dissatisfied=True,
            is_repeated_issue=True,
            repeat_label_expires_at=None,
        ),
        SimpleNamespace(
            call_time=datetime(2026, 6, 19, 9),
            demand_category="操作辅导类",
            topic_category="个税汇算",
            work_order=False,
            resolved_status=True,
            taxpayer_dissatisfied=False,
            is_repeated_issue=False,
            repeat_label_expires_at=None,
        ),
        SimpleNamespace(
            call_time=datetime(2026, 6, 16, 9),
            demand_category="涉税查询类",
            topic_category="其他",
            work_order=True,
            resolved_status=False,
            taxpayer_dissatisfied=False,
            is_repeated_issue=True,
            repeat_label_expires_at=None,
        ),
    ]

    assert _recent_workday_statistics(trajectories) == {
        "start_date": "2026-06-15",
        "end_date": "2026-06-22",
        "call_count": 2,
        "repeated_issue_count": 1,
        "work_order_count": 1,
        "wait_pushback_count": 0,
        "abnormal_end_count": 0,
        "contact_unresolved_count": 0,
        "unresolved_count": 1,
        "dissatisfaction_count": 0,
    }
